import os
from flask import Flask, jsonify, request
from openpyxl import load_workbook
from PIL import Image

app = Flask(__name__)

@app.route('/')
def index():
    return 'HOME'

@app.route('/excel/info', methods=['POST'])
def excel_info():
    sheet = request.files['file']
    sheet = load_workbook(sheet, read_only=True)

    sheetnames = sorted(sheet.sheetnames)
    return jsonify(sheetnames)

@app.route('/image/convert', methods=['POST'])
def image_convert():
    format = request.form['format']
    file = request.files['file']

    filename = os.path.splitext(file.filename)[0]

    with Image.open(file) as image:
        print(filename)
        image.save('{}.{}'.format(filename, format))

    return 'DONE'

if __name__ == '__main__':
    app.run(host='localhost', port=5000)
