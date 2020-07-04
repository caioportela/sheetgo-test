from io import BytesIO

from dropbox import Dropbox
from flask import Flask, abort, jsonify, request, send_file
from flask_httpauth import HTTPTokenAuth
from jwt import decode
from openpyxl import load_workbook
from PIL import Image
from werkzeug.exceptions import HTTPException

app = Flask(__name__)
auth = HTTPTokenAuth(scheme='Bearer')

# Secret key to decode token
SECRET = 'z6Ct_d2Wy0ZcZZVUYD3beI5ZCSsFrR6-f3ZDyn_MW00'

# Accepted emails for authorization
EMAILS = ['lucas@sheetgo.com', 'mauricio@sheetgo.com', 'rafael@sheetgo.com']

# Accepted image formats
FORMATS = ['jpeg', 'png']

@auth.verify_token
def verify_token(token):
    """Decode and verify the auth token."""

    if not token:
        return False

    token_decoded = decode(token, SECRET, algorithms=['HS256'])
    email = token_decoded['email']

    return email in EMAILS

@app.route('/excel/info', methods=['POST'])
@auth.login_required
def excel_info():
    """Return the tabs from the excel file, ordered alphabetically."""

    # Check if the post request has the file part
    if 'file' not in request.files:
        abort(400, description='No file sent')

    sheet = request.files['file']
    sheet = load_workbook(sheet, read_only=True)

    sheetnames = sorted(sheet.sheetnames)
    return jsonify(sheetnames)

@app.route('/image/convert', methods=['POST'])
@auth.login_required
def image_convert():
    """Convert the format of an image."""

    format = request.form.get('format')

    # Check if the format was sent in the request
    if format is None:
        abort(400, description='No format sent')

    # Check if the format sent is accepted
    if format.lower() not in FORMATS:
        abort(400, description='The format sent is invalid')

    # Check if the request has the file part
    if 'file' not in request.files:
        abort(400, description='No file sent')

    file = request.files['file']

    # Create a byte object to save the image
    file_object = BytesIO()

    with Image.open(file) as image:
        image.save(file_object, format)
        file_object.seek(0)

        return send_file(file_object, mimetype='image/{}'.format(format))

@app.route('/image/convert/fromdropbox', methods=['POST'])
@auth.login_required
def convert_from_dropbox():
    """Convert an image from dropbox."""

    access_token = request.headers.get('Dropbox-Token')

    # Check if the dropbox access token was sent
    if access_token is None:
        abort(403, description='No dropbox access token sent')

    dbx = Dropbox(access_token)

    path = request.form.get('path')

    # Check if the path was sent in the request
    if path is None:
        abort(400, description='No path sent')

    format = request.form.get('format')

    # Check if the format was sent in the request
    if format is None:
        abort(400, description='No format sent')

    # Check if the format sent is accepted
    if format.lower() not in FORMATS:
        abort(400, description='The format sent is invalid')

    try:
        metadata, response = dbx.files_download(path)
    except:
        abort(404, description='The file was not found')

    # Create a byte object to save the image
    file = BytesIO()

    with Image.open(BytesIO(response.content)) as image:
        image.save(file, format)
        file.seek(0)

        return send_file(file, mimetype='image/{}'.format(format))

@app.errorhandler(HTTPException)
def handle_exception(e):
    """Handle exception responses."""

    description = '{} {} - {}\n'.format(e.code, e.name, e.description)
    return (description, e.code)

if __name__ == '__main__':
    app.run(host='localhost', port=5000)
